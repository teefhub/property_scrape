import requests
from bs4 import BeautifulSoup
import pandas as pd
import time
import numpy as np
import os
from openpyxl import load_workbook
import seaborn as sns
import re
import pgeocode

pageNumber = 274
def get_suburb_details():
    data = pgeocode.Nominatim('AU')
    details = data.query_postal_code("2117")
    return suburb
def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False, 
                       **to_excel_kwargs):
    # Excel file doesn't exist - saving and exiting
    if not os.path.isfile(filename):
        df.to_excel(
            filename,
            sheet_name=sheet_name, 
            startrow=startrow if startrow is not None else 0, 
            **to_excel_kwargs)
        return
    
    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl', mode='a')

    # try to open an existing workbook
    writer.book = load_workbook(filename)
    
    # get the last row in the existing Excel sheet
    # if it was not specified explicitly
    if startrow is None and sheet_name in writer.book.sheetnames:
        startrow = writer.book[sheet_name].max_row

    # truncate sheet
    if truncate_sheet and sheet_name in writer.book.sheetnames:
        # index of [sheet_name] sheet
        idx = writer.book.sheetnames.index(sheet_name)
        # remove [sheet_name]
        writer.book.remove(writer.book.worksheets[idx])
        # create an empty sheet [sheet_name] using old index
        writer.book.create_sheet(sheet_name, idx)
    
    # copy existing sheets
    writer.sheets = {ws.title:ws for ws in writer.book.worksheets}

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()
def get_property():
    page = 1
    ls= []
    url = 'https://www.auhouseprices.com/sold/list/NSW/2155/Rouse+Hill/'
    for page in range(1, pageNumber):
        response = requests.get(url=url+str(page))
        c = response.content
        soup = BeautifulSoup(c, "html.parser")
        all_c = soup.findAll("div", {"class": "caption"})
        for item in all_c:
            d={}
            d["price"]=item.find("span",{"class":"pull-right"}).text
            d["address"]=item.find("h4").text
            d["date"] = re.findall(r"\d{1,2}\s\w+\s\d{4}",item.find("ul",{"class":"list-unstyled"}).text)
            d["bed/bath/car"] = re.findall(r"\d{1}\s\d{1}\s\d{1}",item.find("ul",{"class":"list-unstyled"}).text)
            d["land size"] = re.findall(r"(?<=Land Size:).\w+\s\w+",item.find("ul",{"class":"list-unstyled"}).text)
            ls.append(d)
        page=+1
        
    df = pd.DataFrame(ls)
    return df
#timeD=time.strftime("%Y-%m-%d-%H-%S",time.localtime())
append_df_to_excel("2021-09-06-22-25-get_previous_price.xlsx", get_property(),"rouse hill")
#sns.countplot(get_property()['price'])